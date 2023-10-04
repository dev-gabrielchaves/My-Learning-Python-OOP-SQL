import openpyxl

# Creating a spreadsheet(book)
book = openpyxl.Workbook()
# How to visualize existent pages
print(book.sheetnames)
# How to create a new sheet to the existent spreadsheet
book.create_sheet('Fruits')
# How to select and add items to a sheet
fruits_page = book['Fruits']
fruits_page.append(['Fruit Name', 'Amount', 'Price'])
fruits_page.append(['Banana', '5', '$3,90'])
fruits_page.append(['Apple', '2', '$1,90'])
fruits_page.append(['Pear', '7', '$7,90'])
# How to save the spreadsheet
book.save('Buying_List.xlsx')