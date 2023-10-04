import openpyxl

wb = openpyxl.Workbook()

# A workbook is always created with at least one worksheet. You can get it by using the Workbook.active property:
ws1 = wb.active
# This is set to 0 by default. Unless you modify its value, you will always get the first worksheet by using this method.

# Setting the title of our first worksheet.
ws1.title = 'Default'

# To create another sheets is as simple as...
ws2 = wb.create_sheet('Another sheet')

# Creating a loop just to define some elements in our cells
for row in range(1, 11):
    for column in range(0, 10):
        char = chr(65 + column)
        ws1[char + str(row)] = chr(65 + column) + str(row)

# To merge cells you do this:
ws1.merge_cells('A1:E5')

# And to unmerge you do this:
ws1.unmerge_cells('A1:E5')
# If you merge and unmerge after you can't get back the data that you had.

wb.save('any_workbook.xlsx')